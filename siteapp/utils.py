from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_queryset_to_xlsx(queryset, filename, headers, row_builder):
    wb = Workbook(); ws = wb.active; ws.title = 'Данные'; ws.append(headers)
    fill = PatternFill('solid', fgColor='E8F1DD')
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = fill; cell.alignment = Alignment(horizontal='center')
    for obj in queryset: ws.append(row_builder(obj))
    for col in ws.columns:
        width = max(len(str(c.value or '')) for c in col) + 3
        ws.column_dimensions[col[0].column_letter].width = min(width, 60)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_queryset_to_html(queryset, filename, title, headers, row_builder):
    rows_html = []
    for item in queryset:
        row = ''.join(f'<td>{value if value is not None else ""}</td>' for value in row_builder(item))
        rows_html.append(f'<tr>{row}</tr>')

    header_html = ''.join(f'<th>{header}</th>' for header in headers)
    html = f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>{title}</title>
  <style>
    body {{ font-family: Arial, sans-serif; padding: 24px; color: #2f261f; }}
    h1 {{ color: #5a4636; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #d8c9b5; padding: 8px 10px; vertical-align: top; }}
    th {{ background: #f6efe4; }}
    tr:nth-child(even) {{ background: #fcfaf6; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <table>
    <thead><tr>{header_html}</tr></thead>
    <tbody>{''.join(rows_html)}</tbody>
  </table>
</body>
</html>"""
    response = HttpResponse(html, content_type='text/html; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_queryset_to_csv(queryset, filename, headers, row_builder):
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    for obj in queryset:
        writer.writerow(row_builder(obj))
    return response
